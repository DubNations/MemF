from __future__ import annotations

import io
import zipfile
from typing import Any, Dict, List, Optional
from xml.etree import ElementTree

from cognitive_os.ingestion.parsers.base_parser import (
    BaseParser,
    FileExtension,
    ParseResult,
    ParseStrategy,
    TableData,
)


class MegaparseAdapter(BaseParser):
    supported_extensions = [
        FileExtension.PDF,
        FileExtension.DOC,
        FileExtension.DOCX,
        FileExtension.PPTX,
        FileExtension.TXT,
        FileExtension.MD,
        FileExtension.CSV,
        FileExtension.XLSX,
    ]

    def __init__(
        self,
        enable_ocr: bool = True,
        enable_table_extraction: bool = True,
        ocr_language: str = "chi_sim+eng",
    ):
        self.enable_ocr = enable_ocr
        self.enable_table_extraction = enable_table_extraction
        self.ocr_language = ocr_language
        self._megaparse_available = self._check_megaparse()
        self._tesseract_available = self._check_tesseract()
        self._pdfplumber_available = self._check_pdfplumber()

    def _check_megaparse(self) -> bool:
        try:
            import megaparse
            return True
        except ImportError:
            return False

    def _check_tesseract(self) -> bool:
        try:
            import pytesseract
            return True
        except ImportError:
            return False

    def _check_pdfplumber(self) -> bool:
        try:
            import pdfplumber
            return True
        except ImportError:
            return False

    def parse(
        self,
        content: bytes,
        filename: str,
        strategy: ParseStrategy = ParseStrategy.AUTO,
        **kwargs: Any,
    ) -> ParseResult:
        ext = self.get_extension(filename)

        if self._megaparse_available:
            return self._parse_with_megaparse(content, filename, strategy, **kwargs)

        if ext == ".pdf":
            return self._parse_pdf_enhanced(content, filename, strategy)
        elif ext == ".docx":
            return self._parse_docx_enhanced(content, filename)
        elif ext == ".pptx":
            return self._parse_pptx(content, filename)
        elif ext == ".xlsx":
            return self._parse_xlsx(content, filename)
        elif ext in (".txt", ".md", ".csv"):
            return self._parse_text(content, filename)
        else:
            return ParseResult(
                text="",
                error=f"Unsupported format: {ext}",
                metadata={"filename": filename},
            )

    def parse_from_path(
        self,
        file_path: str,
        strategy: ParseStrategy = ParseStrategy.AUTO,
        **kwargs: Any,
    ) -> ParseResult:
        try:
            with open(file_path, "rb") as f:
                content = f.read()
            filename = file_path.split("/")[-1].split("\\")[-1]
            return self.parse(content, filename, strategy, **kwargs)
        except Exception as e:
            return ParseResult(
                text="",
                error=str(e),
                metadata={"file_path": file_path},
            )

    def _parse_with_megaparse(
        self,
        content: bytes,
        filename: str,
        strategy: ParseStrategy = ParseStrategy.AUTO,
        **kwargs: Any,
    ) -> ParseResult:
        try:
            from megaparse import MegaParse
            from megaparse_sdk.schema.parser_config import StrategyEnum as MPStrategy

            strategy_map = {
                ParseStrategy.AUTO: MPStrategy.AUTO,
                ParseStrategy.FAST: MPStrategy.FAST,
                ParseStrategy.HI_RES: MPStrategy.HI_RES,
            }

            megaparse = MegaParse()
            mp_strategy = strategy_map.get(strategy, MPStrategy.AUTO)

            import tempfile
            with tempfile.NamedTemporaryFile(delete=False, suffix=self.get_extension(filename)) as tmp:
                tmp.write(content)
                tmp_path = tmp.name

            result = megaparse.convert(tmp_path, strategy=mp_strategy)

            tables = []
            if self.enable_table_extraction and hasattr(result, "tables"):
                for table in result.tables:
                    tables.append(TableData(
                        headers=table.get("headers", []),
                        rows=table.get("rows", []),
                        caption=table.get("caption", ""),
                    ))

            text = str(result) if result else ""

            return ParseResult(
                text=text,
                tables=tables,
                pages=result.pages if hasattr(result, "pages") else 1,
                strategy_used=strategy,
                ocr_used=mp_strategy == MPStrategy.HI_RES,
                metadata={"filename": filename, "parser": "megaparse"},
            )
        except Exception as e:
            return self._parse_fallback(content, filename, strategy, str(e))

    def _parse_fallback(
        self,
        content: bytes,
        filename: str,
        strategy: ParseStrategy,
        error: str,
    ) -> ParseResult:
        from cognitive_os.ingestion.parsers.native_parser import NativeParser
        native = NativeParser()
        result = native.parse(content, filename, strategy)
        result.metadata["fallback_used"] = True
        result.metadata["original_error"] = error
        return result

    def _parse_pdf_enhanced(
        self,
        content: bytes,
        filename: str,
        strategy: ParseStrategy = ParseStrategy.AUTO,
    ) -> ParseResult:
        text_parts: List[str] = []
        tables: List[TableData] = []
        pages = 1
        ocr_used = False

        if self._pdfplumber_available:
            result = self._parse_with_pdfplumber(content, filename)
            text_parts.append(result.get("text", ""))
            tables.extend(result.get("tables", []))
            pages = result.get("pages", 1)

        if not text_parts or not "".join(text_parts).strip():
            if self.enable_ocr and self._tesseract_available:
                ocr_result = self._parse_with_ocr(content, filename)
                text_parts.append(ocr_result.get("text", ""))
                ocr_used = True

        if not text_parts:
            from cognitive_os.ingestion.parsers.native_parser import NativeParser
            native = NativeParser()
            native_result = native.parse(content, filename, strategy)
            text_parts.append(native_result.text)
            pages = native_result.pages

        final_text = "\n\n".join([t for t in text_parts if t.strip()])

        return ParseResult(
            text=final_text,
            tables=tables,
            pages=pages,
            strategy_used=ParseStrategy.HI_RES if ocr_used else ParseStrategy.FAST,
            ocr_used=ocr_used,
            metadata={"filename": filename, "parser": "enhanced_native"},
        )

    def _parse_with_pdfplumber(self, content: bytes, filename: str) -> Dict[str, Any]:
        import pdfplumber

        text_parts = []
        tables = []
        pages = 0

        with pdfplumber.open(io.BytesIO(content)) as pdf:
            pages = len(pdf.pages)
            for page in pdf.pages:
                page_text = page.extract_text() or ""
                if page_text.strip():
                    text_parts.append(page_text)

                if self.enable_table_extraction:
                    page_tables = page.extract_tables()
                    for table in page_tables:
                        if table and len(table) > 1:
                            headers = [str(cell or "") for cell in table[0]]
                            rows = [[str(cell or "") for cell in row] for row in table[1:]]
                            tables.append(TableData(headers=headers, rows=rows))

        return {"text": "\n\n".join(text_parts), "tables": tables, "pages": pages}

    def _parse_with_ocr(self, content: bytes, filename: str) -> Dict[str, Any]:
        import pytesseract
        from pdf2image import convert_from_bytes

        images = convert_from_bytes(content)
        text_parts = []

        for img in images:
            text = pytesseract.image_to_string(img, lang=self.ocr_language)
            if text.strip():
                text_parts.append(text)

        return {"text": "\n\n".join(text_parts)}

    def _parse_docx_enhanced(self, content: bytes, filename: str) -> ParseResult:
        text_parts: List[str] = []
        tables: List[TableData] = []

        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                doc_data = zf.read("word/document.xml")
                root = ElementTree.fromstring(doc_data)

                text_nodes = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
                text_parts.append("\n".join(text_nodes))

                for tbl in root.iter():
                    if tbl.tag.endswith("}tbl"):
                        rows = []
                        for row in tbl.iter():
                            if row.tag.endswith("}tr"):
                                cells = [cell.text for cell in row.iter() if cell.tag.endswith("}t") and cell.text]
                                if cells:
                                    rows.append(cells)

                        if rows:
                            headers = rows[0] if rows else []
                            table_rows = rows[1:] if len(rows) > 1 else []
                            tables.append(TableData(headers=headers, rows=table_rows))

            return ParseResult(
                text="\n\n".join(text_parts),
                tables=tables,
                pages=1,
                strategy_used=ParseStrategy.FAST,
                metadata={"filename": filename},
            )
        except Exception as e:
            return ParseResult(
                text="",
                error=str(e),
                metadata={"filename": filename},
            )

    def _parse_pptx(self, content: bytes, filename: str) -> ParseResult:
        text_parts: List[str] = []
        slide_count = 0

        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                slide_files = sorted([n for n in zf.namelist() if n.startswith("ppt/slides/slide") and n.endswith(".xml")])
                slide_count = len(slide_files)

                for slide_file in slide_files:
                    slide_data = zf.read(slide_file)
                    root = ElementTree.fromstring(slide_data)
                    text_nodes = [node.text for node in root.iter() if node.tag.endswith("}t") and node.text]
                    if text_nodes:
                        text_parts.append(f"[Slide {slide_files.index(slide_file) + 1}]\n" + "\n".join(text_nodes))

            return ParseResult(
                text="\n\n".join(text_parts),
                pages=slide_count,
                strategy_used=ParseStrategy.FAST,
                metadata={"filename": filename, "slide_count": slide_count},
            )
        except Exception as e:
            return ParseResult(
                text="",
                error=str(e),
                metadata={"filename": filename},
            )

    def _parse_xlsx(self, content: bytes, filename: str) -> ParseResult:
        tables: List[TableData] = []
        text_parts: List[str] = []

        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))

            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                rows = []
                for row in sheet.iter_rows(values_only=True):
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(row_data):
                        rows.append(row_data)

                if rows:
                    headers = rows[0] if rows else []
                    table_rows = rows[1:] if len(rows) > 1 else []
                    tables.append(TableData(
                        headers=headers,
                        rows=table_rows,
                        caption=f"Sheet: {sheet_name}",
                    ))
                    text_parts.append(f"[Sheet: {sheet_name}]\n{tables[-1].to_markdown()}")

            return ParseResult(
                text="\n\n".join(text_parts),
                tables=tables,
                pages=len(tables),
                strategy_used=ParseStrategy.FAST,
                metadata={"filename": filename, "sheet_count": len(wb.sheetnames)},
            )
        except ImportError:
            return ParseResult(
                text="",
                error="openpyxl not installed, cannot parse xlsx",
                metadata={"filename": filename},
            )
        except Exception as e:
            return ParseResult(
                text="",
                error=str(e),
                metadata={"filename": filename},
            )

    def _parse_text(self, content: bytes, filename: str) -> ParseResult:
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError:
            text = content.decode("latin-1", errors="ignore")

        return ParseResult(
            text=text,
            pages=len(text.split("\n\n")),
            strategy_used=ParseStrategy.FAST,
            metadata={"filename": filename},
        )
